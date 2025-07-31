import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook

st.set_page_config(
    page_title="SKU Aggregator - Maestro con Distribución Global",
    page_icon="📦",
    layout="wide",
)

st.title("📦 SKU Aggregator - Maestro con Distribución de Paletas Global")
st.markdown(
    """
    1. Sube tu **archivo maestro** (Excel) con formato y colores propios.  
    2. Sube **Vitaplena.xlsx** (columna D desde fila 2).  
    3. Sube **Eggmarket.xlsx** (columna F desde fila 7).  
    4. La app sumará todas las apariciones de cada SKU.  
    5. Actualiza la columna **master** (header en O3) con totales.  
    6. Distribuye estos totales en paletas globales (columnas B–M) con capacidad total de 30 bultos por paleta, combinando SKUs.
    """
)

# Uploaders
t1, t2, t3 = st.columns([1,1,1])
with t1:
    master_file = st.file_uploader("1️⃣ Sube tu archivo maestro (xlsx)", type=["xlsx"], key="master")
with t2:
    vitaplena_file = st.file_uploader("2️⃣ Sube Vitaplena.xlsx", type=["xlsx","xls"], key="vita")
with t3:
    egg_file = st.file_uploader("3️⃣ Sube Eggmarket.xlsx", type=["xlsx","xls"], key="egg")

# Validar cargas
if not master_file:
    st.info("Por favor, sube primero el archivo maestro.")
elif not (vitaplena_file and egg_file):
    st.info("Ahora sube Vitaplena.xlsx y Eggmarket.xlsx.")
else:
    # Cargar maestro con openpyxl
    wb = load_workbook(filename=master_file, data_only=False)
    ws = wb.active

    # Leer SKUs de Vitaplena y Eggmarket
    df_vita = pd.read_excel(vitaplena_file, usecols=[3], skiprows=1, names=["SKU"])
    df_egg = pd.read_excel(egg_file, usecols=[5], skiprows=6, names=["SKU"])
    df_vita["SKU"] = df_vita["SKU"].astype(str).apply(lambda x: x.split(':',1)[-1] if ':' in x else x)
    df_egg["SKU"] = df_egg["SKU"].astype(str).apply(lambda x: x.split(':',1)[-1] if ':' in x else x)

    # Unir y contar totales por SKU
    counts = pd.concat([df_vita, df_egg], ignore_index=True)
    summary = counts["SKU"].value_counts().rename_axis("SKU").reset_index(name="Total")
    summary["Total"] = summary["Total"].astype(int)

    # Identificar columna 'master' (fila 3)
    master_col = None
    for col in range(1, ws.max_column+1):
        if str(ws.cell(row=3, column=col).value).strip().lower() == 'master':
            master_col = col
            break
    if master_col is None:
        st.error("No se encontró la columna 'master' en la fila 3.")
        st.stop()

    # Escribir totales y distribuir globalmente en paletas B–M
    pallet_cols = list(range(2,14))  # B=2,...,M=13
    MAX_CAPACITY = 30
    # Crear map de totales por SKU
    total_map = dict(zip(summary['SKU'], summary['Total']))
    # Variables de control global
    pallet_idx = 0
    used_in_current = 0

    # Recorrer SKUs en maestro: columna A fila >=4
    for row in range(4, ws.max_row+1):
        sku_val = ws.cell(row=row, column=1).value
        if sku_val is None:
            continue
        sku_key = str(sku_val).split(':',1)[-1] if ':' in str(sku_val) else str(sku_val)
        total_qty = total_map.get(sku_key, 0)
        # Escribir total en 'master' columna
        ws.cell(row=row, column=master_col, value=total_qty)
        # Distribuir global
        remaining = total_qty
        while remaining > 0 and pallet_idx < len(pallet_cols):
            space_left = MAX_CAPACITY - used_in_current
            to_assign = min(space_left, remaining)
            # Acumular en la paleta actual
            current_cell = ws.cell(row=row, column=pallet_cols[pallet_idx])
            existing = current_cell.value or 0
            current_cell.value = existing + to_assign
            remaining -= to_assign
            used_in_current += to_assign
            # Si paleta llena, pasar a siguiente
            if used_in_current == MAX_CAPACITY:
                pallet_idx += 1
                used_in_current = 0
        # Si se agotan paletas, detenemos asignación

    # Preview primeras líneas
    preview = []
    headers = ['SKU','Total'] + [f'Pal{pallet_idx+1}' for pallet_idx in range(len(pallet_cols))]
    for row in range(4, min(ws.max_row, 13)+1):
        row_vals = [
            ws.cell(row=row, column=1).value,
            ws.cell(row=row, column=master_col).value
        ] + [ws.cell(row=row, column=col).value for col in pallet_cols]
        preview.append(row_vals)
    st.subheader("✅ Distribución Global de Paletas (filas 4-13)")
    st.dataframe(pd.DataFrame(preview, columns=headers), use_container_width=True)

    # Descargar
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    st.download_button(
        "📥 Descargar Maestro con Totales y Paletas",
        data=output,
        file_name="maestro_totales_paletas_global.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
